import datetime
import os
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.core.paginator import Paginator
from django.db.models import Q, F, Subquery, OuterRef, Count, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse, Http404, HttpResponseNotFound
from django.shortcuts import render, redirect
from openpyxl.reader.excel import load_workbook

from apps.dashboard.filters import ProductFilter, PromoFilter, CashbekFilter, SellerFilter, PaymentSeller, PaymentVendor
from apps.main.models import *
from apps.main.utility import get_tashkent_time
from config.settings import BASE_DIR


def dashboard_access(user):
    if user.is_authenticated and user.is_dashboard():
        return True
    return False


def signin(request):
    if request.method == "POST":
        phone = request.POST.get("phone")
        password = request.POST.get("password")
        remember = request.POST.get("remember")
        user = authenticate(username=phone, password=password)
        if user and user.is_dashboard():
            login(request, user)
            if not remember:
                request.session.set_expiry(0)
                request.session.modified = True
            return redirect("index")
        else:
            messages.error(request, "Invalit login or password")
    return render(request, "login.html")


@user_passes_test(dashboard_access, login_url="signin")
def index(request):
    context = {"home": True}
    if request.user.is_manager():
        cashbek = Cashbek.objects.filter(active=True)
        f = CashbekFilter(request.GET, queryset=cashbek)
        cashbek = f.qs
        context.update({"filter": f})
    else:
        cashbek = Cashbek.objects.filter(active=True, vendor=request.user.vendor.vendor)
    incom_count = cashbek.filter(types=1).count()
    expen_count = cashbek.filter(types=2).count()
    shop = cashbek.values("seller__name").distinct().count()
    all_cashbek = cashbek.annotate(day=TruncDate('created_at')).values('day').annotate(count=Count('id'))
    cashbek_count_list = []
    incom_count_list = []
    expen_count_list = []
    datetime_list = []
    for i in all_cashbek:
        datetime_list.append(i["day"].strftime("%Y-%m-%d"))
        cashbek_count_list.append(i["count"])
        incom_count_list.append(cashbek.filter(types=1, created_at__date=i["day"]).count())
        expen_count_list.append(cashbek.filter(types=2, created_at__date=i["day"]).count())
    shop_names = []
    shop_count_list = []
    for i in cashbek.values('seller__name').annotate(seller_count=Count('seller')).order_by('-seller_count')[:10]:
        shop_names.append(i["seller__name"])
        shop_count_list.append(i["seller_count"])

    context.update({"cashbeks": cashbek, "incom_count": incom_count, "expen_count": expen_count, "shops": shop,
                    "datetime_list": datetime_list, "cashbek_count_list": cashbek_count_list,
                    "incom_count_list": incom_count_list, "expen_count_list": expen_count_list,
                    "shop_names": shop_names, "shop_count_list": shop_count_list})
    return render(request, "index.html", context)


@user_passes_test(dashboard_access, login_url="signin")
def products(request):
    context = {"products": True}
    if request.user.is_manager():
        products = Products.objects.all().order_by("-id")
        alls = products
        active = products.filter(is_active=True)
        no_active = products.filter(is_active=False)
        f = ProductFilter(request.GET, queryset=products)
        products = f.qs.order_by("-id")
        context.update({"filter": f})
    else:
        products = Products.objects.filter(ven=request.user.vendor.vendor).order_by("-id")
        alls = products
        active = products.filter(is_active=True)
        no_active = products.filter(is_active=False)
    if request.GET.get('q'):
        word = request.GET.get('q')
        products = products.filter(Q(model__icontains=word) | Q(imei1__icontains=word) | Q(sku__icontains=word))
    pagination = Paginator(products, 50)
    page_number = request.GET.get('page')
    page_obj = pagination.get_page(page_number)
    context.update({"objs": page_obj, "all_products": products, "all": alls, "active": active, "no_active": no_active})
    return render(request, "products.html", context)


@user_passes_test(dashboard_access, login_url="signin")
def export_example(request, type):
    if type == "products":
        file_location = BASE_DIR / 'file/example_products.xlsx'
    else:
        file_location = BASE_DIR / 'file/example_promo.xlsx'
    with open(file_location, 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    return response


@user_passes_test(dashboard_access, login_url="signin")
def import_products(request):
    file = request.FILES['file']
    filename = file.name
    if filename.split(".")[1] != "xlsx":
        messages.error(request, "Файл должен быть в формате .xlsx")
    else:
        products = Products.objects.all()
        error_import = []
        creator = []
        imei = []
        wb = load_workbook(file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, max_col=3):
            data = {"model": "", "sku": "", "imei1": "", "reason": ""}
            if row[0].value is None or row[1].value is None or row[2].value is None:
                data["model"] = row[0].value
                data["imei1"] = row[1].value
                data["sku"] = row[2].value
                data["reason"] = "Обязательные строки не заполняются"
                error_import.append(data)
                continue
            elif products.filter(imei1=row[1].value):
                data["model"] = row[0].value
                data["imei1"] = row[1].value
                data["sku"] = row[2].value
                data["reason"] = "Этот продукт указан"
                error_import.append(data)
                continue
            elif row[1].value in imei:
                data["model"] = row[0].value
                data["imei1"] = row[1].value
                data["sku"] = row[2].value
                data["reason"] = "Этот продукт указан"
                error_import.append(data)
                continue
            creator.append(
                BlackListProducts(model=row[0].value, imei1=row[1].value, sku=row[2].value, vendor=request.user.vendor))
            imei.append(row[1].value)
        context = {"products": True, "error": error_import, "correct": creator}
        if creator:
            obj = BlackListProducts.objects.bulk_create(creator)
            context["interval"] = f"{obj[0].id}-{obj[-1].id}"
        return render(request, "confirm_products.html", context)

    return redirect("products")


@user_passes_test(dashboard_access, login_url="signin")
def export_products(request):
    if request.user.is_manager():
        products = Products.objects.all().order_by("-id")
    else:
        products = Products.objects.filter(ven=request.user.vendor.vendor).order_by("-id")
    file_location = BASE_DIR / 'file/products.xlsx'
    file_send = BASE_DIR / f'file/products_{timezone.now().strftime("%d-%m-%Y")}.xlsx'
    wb = load_workbook(file_location)
    sheet = wb.active
    for i_row, user_data in enumerate(products, start=2):
        sheet.cell(row=i_row, column=1, value=user_data.model)
        sheet.cell(row=i_row, column=2, value=user_data.datetime.strftime("%d-%m-%Y"))
        sheet.cell(row=i_row, column=3, value=user_data.imei1)
        sheet.cell(row=i_row, column=4, value=user_data.sku)
        sheet.cell(row=i_row, column=5, value="Актив" if user_data.is_active else "Не актив")
    wb.save(file_send)
    with open(file_send, 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="products_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'
    os.remove(file_send)
    return response


@user_passes_test(dashboard_access, login_url="signin")
def confirm_products(request):
    if request.POST.get("new_prs"):
        prs = request.POST.get("new_prs")
        prs = prs.split("-")
        creator = []
        objs = BlackListProducts.objects.filter(pk__gte=prs[0], pk__lte=prs[1], vendor=request.user.vendor)
        for obj in objs:
            creator.append(Products(model=obj.model, imei1=obj.imei1, sku=obj.sku, vendor=request.user.vendor,
                                    ven=request.user.vendor.vendor))
        Products.objects.bulk_create(creator)
        objs.delete()
        messages.success(request, "Товары успешно добавлены")
        return redirect("products")


@user_passes_test(dashboard_access, login_url="signin")
def promo(request):
    context = {"promo": True}
    if request.user.is_manager():
        promos = Promo.objects.all().order_by("-id")
        f = PromoFilter(request.GET, queryset=promos)
        promos = f.qs.order_by("-id")
        context.update({"filter": f})
    else:
        promos = Promo.objects.filter(ven=request.user.vendor.vendor).order_by("-id")
    if request.GET.get('q'):
        word = request.GET.get('q')
        promos = promos.filter(name__icontains=word)
    pagination = Paginator(promos, 10)
    page_number = request.GET.get('page')
    page_obj = pagination.get_page(page_number)
    context.update({"obj": page_obj})
    return render(request, "promo.html", context)


@user_passes_test(dashboard_access, login_url="signin")
def import_promo(request):
    try:
        file = request.FILES['file']
    except:
        return redirect("promo")
    filename = file.name
    name = request.POST.get("name")
    now = datetime.now().date()
    start_promo = datetime.strptime(request.POST.get("start_promo"), "%Y-%m-%d").date()
    end_promo = datetime.strptime(request.POST.get("end_promo"), "%Y-%m-%d").date()
    budget = request.POST.get("budjet")
    if filename.split(".")[1] != "xlsx":
        messages.error(request, "Файл должен быть в формате .xlsx")
    elif end_promo <= start_promo or start_promo < now:
        messages.error(request, "Дата начала должна быть меньше даты окончания, а дата начала должна быть больше "
                                "текущей даты.")
    else:
        TempPromo.objects.filter(vendor=request.user.vendor).delete()
        TempPriceProduct.objects.filter(promo__vendor=request.user.vendor).delete()
        products = Products.objects.all()
        error_import = []
        correct = []
        imei = []
        amount = 0
        wb = load_workbook(file)
        sheet = wb.active
        promo = TempPromo.objects.create(name=name, budget=budget, start=start_promo, end=end_promo,
                                         vendor=request.user.vendor)
        for row in sheet.iter_rows(min_row=2, max_col=3):
            data = {"model": row[0].value, "amount": row[2].value, "imei1": row[1].value, "reason": ""}
            if row[0].value is None or row[1].value is None or row[2].value is None:
                data["reason"] = "Обязательные строки не заполняются"
                error_import.append(data)
                continue
            elif not products.filter(imei1=row[1].value):
                data["reason"] = "Этот продукт недоступен в основной базе"
                error_import.append(data)
                continue
            elif not products.filter(imei1=row[1].value).first().is_active:
                data["reason"] = "Этот продукт не активен"
                error_import.append(data)
                continue
            elif row[1].value in imei:
                data["reason"] = "Этот продукт указан"
                error_import.append(data)
                continue
            elif isinstance(row[2].value, str) and not row[2].value.isnumeric():
                data["reason"] = "Введите цену числом"
                error_import.append(data)
                continue
            elif Promo.objects.filter(Q(products__imei1=row[1].value) & (Q(status=WAIT) | Q(status=ACTIVE))):
                data["reason"] = "Еще один актив доступен по акции"
                error_import.append(data)
                continue
            amount += int(row[2].value)
            correct.append(products.filter(imei1=row[1].value).first())
            TempPriceProduct.objects.create(promo=promo, product=products.filter(imei1=row[1].value).first(),
                                            price=int(row[2].value))
            imei.append(row[1].value)
        context = {"promo": True, "error": error_import, "correct": correct, "promos": promo}

        if correct:
            if int(budget) > amount:
                context["old_budget"] = budget
                promo.budget = amount
                promo.save()
            promo.products.set(correct)
            context["promos"] = promo
            context["product"] = promo.products.annotate(price=Subquery(
                TempPriceProduct.objects.filter(
                    promo=promo,
                    product_id=OuterRef('pk')
                ).values('price')[:1]
            ))
        return render(request, "confirm_promo.html", context)

    return redirect("promo")


@user_passes_test(dashboard_access, login_url="signin")
def confirm_promo(request):
    if request.POST.get("promo"):
        pid = request.POST.get("promo")
        new = TempPromo.objects.get(pk=pid)
        obj = Promo.objects.create(name=new.name, start=new.start, end=new.end, budget=new.budget,
                                   ven=request.user.vendor.vendor,
                                   vendor=request.user.vendor, price_procent=request.user.vendor.vendor.price)
        obj.products.set(new.products.all())
        bulk_obj = []
        for i in TempPriceProduct.objects.filter(promo=new):
            if obj.price_procent == 100:
                amount = i.price
            else:
                amount = round(i.price * obj.price_procent / 100, ndigits=-3)
            bulk_obj.append(PriceProduct(promo=obj, product=i.product, price=amount, all_price=i.price))
        PriceProduct.objects.bulk_create(bulk_obj)
        messages.success(request, "Промо успешно создано")
        return redirect("promo")


@user_passes_test(dashboard_access, login_url="signin")
def export_promo(request, pk):
    promo = Promo.objects.get(pk=pk)
    if not request.user.is_manager() and promo.vendor != request.user.vendor:
        return HttpResponseNotFound()
    file_location = BASE_DIR / 'file/example_romo_products.xlsx'
    file_send = BASE_DIR / f'file/{promo.vendor.vendor}_{promo.name}.xlsx'
    wb = load_workbook(file_location)
    sheet = wb.active
    sheet["C1"].value = promo.name
    sheet["C2"].value = promo.create_at.strftime("%d-%m-%Y")
    sheet["C3"].value = promo.start.strftime("%d-%m-%Y")
    sheet["C4"].value = promo.end.strftime("%d-%m-%Y")
    sheet["C5"].value = promo.budget
    for i_row, user_data in enumerate(promo.products.annotate(price=Subquery(
            PriceProduct.objects.filter(
                promo=promo,
                product_id=OuterRef('pk')
            ).values('price')[:1]
    )), start=8):
        sheet.cell(row=i_row, column=1, value=user_data.model)
        sheet.cell(row=i_row, column=2, value=user_data.imei1)
        sheet.cell(row=i_row, column=3, value=user_data.sku)
        sheet.cell(row=i_row, column=4, value=user_data.price)
        sheet.cell(row=i_row, column=5, value="Актив" if user_data.is_active else "Не актив")
    wb.save(file_send)
    with open(file_send, 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{promo.vendor.vendor.name}_{promo.name}.xlsx"'
    os.remove(file_send)
    return response


@user_passes_test(dashboard_access, login_url="signin")
def confirm_status(request):
    promo = Promo.objects.get(pk=request.POST.get("promo"))
    if request.POST.get("confirm"):
        promo.status = ACTIVE
        promo.description = request.POST.get("comment")
        promo.who = request.POST.get("who")
        promo.save()
    elif request.POST.get("refuse"):
        promo.status = REFUSED
        promo.description = request.POST.get("comment")
        promo.who = request.POST.get("who")
        promo.save()
    elif request.POST.get("pause"):
        promo.status = PAUSE
        promo.description = request.POST.get("comment")
        promo.who = request.POST.get("who")
        promo.save()
    elif request.POST.get("continue"):
        promo.status = ACTIVE
        promo.description = request.POST.get("comment")
        promo.who = request.POST.get("who")
        promo.save()
    elif request.POST.get("finish"):
        promo.status = FINISH
        promo.description = request.POST.get("comment")
        promo.who = request.POST.get("who")
        promo.save()
    return redirect("promo")


@user_passes_test(dashboard_access, login_url="signin")
def cashbek(request):
    context = {"cashbek": True}
    if request.user.is_manager():
        cashbek = Cashbek.objects.all().order_by("-id")
        f = CashbekFilter(request.GET, queryset=cashbek)
        cashbek = f.qs.order_by("-id")
        incom = cashbek.filter(active=True, types=1).aggregate(all_price=Sum(F("price")))["all_price"]
        excom = cashbek.filter(active=True, types=2).aggregate(all_price=Sum(F("price")))["all_price"]
        context.update({"filter": f, 'incom': incom, 'excom': excom})
    else:
        cashbek = Cashbek.objects.filter(vendor=request.user.vendor.vendor, active=True).order_by("-id")
        f = CashbekFilter(request.GET, queryset=cashbek)
        cashbek = f.qs.order_by("-id")
        incom = cashbek.filter(types=1).aggregate(all_price=Sum(F("price")))["all_price"]
        excom = cashbek.filter(types=2).aggregate(all_price=Sum(F("price")))["all_price"]
        context.update({"filter": f, 'incom': incom, 'excom': excom})
    if request.GET.get('start_date'):
        context.update({"start_date": request.GET.get('start_date')})
    if request.GET.get('end_date'):
        context.update({"end_date": request.GET.get('end_date')})
    if request.GET.get('vendor'):
        context.update({"vendor_id": request.GET.get('vendor')})
    if request.GET.get('q'):
        word = request.GET.get('q')
        if request.user.is_manager():
            cashbek = cashbek.filter(Q(product__model__icontains=word) | Q(product__imei1__icontains=word) | Q(
                promo__name__icontains=word) | Q(user_phone__icontains=word) | Q(seller__name__icontains=word) | Q(
                user__last_name__icontains=word))
            incom = cashbek.filter(active=True, types=1).aggregate(all_price=Sum(F("price")))["all_price"]
            excom = cashbek.filter(active=True, types=2).aggregate(all_price=Sum(F("price")))["all_price"]
            context.update({'incom': incom, 'excom': excom})
        else:
            cashbek = cashbek.filter(Q(product__model__icontains=word) | Q(product__imei1__icontains=word) | Q(
                promo__name__icontains=word) | Q(seller__name__icontains=word))
            incom = cashbek.filter(types=1).aggregate(all_price=Sum(F("price")))["all_price"]
            excom = cashbek.filter(types=2).aggregate(all_price=Sum(F("price")))["all_price"]
            context.update({'incom': incom, 'excom': excom})
    pagination = Paginator(cashbek, 10)
    page_number = request.GET.get('page')
    page_obj = pagination.get_page(page_number)
    context.update({"obj": page_obj})
    return render(request, "cashbek.html", context)


@user_passes_test(dashboard_access, login_url="signin")
def shop(request):
    context = {"shop": True}
    if request.user.is_manager():
        cashbek = Cashbek.objects.filter(active=True)
        seller = Seller.objects.filter(cash_seller__active=True).distinct()
    else:
        cashbek = Cashbek.objects.filter(vendor=request.user.vendor.vendor, active=True)
        seller = Seller.objects.filter(cash_seller__active=True,
                                       cash_seller__vendor=request.user.vendor.vendor).distinct()
    if request.GET.get('start_date'):
        context.update({"start_date": request.GET.get('start_date')})
    if request.GET.get('end_date'):
        context.update({"end_date": request.GET.get('end_date')})
    if request.GET.get('q'):
        word = request.GET.get('q')
        if request.user.is_manager():
            seller = Seller.objects.filter(cash_seller__active=True).distinct()
        else:
            seller = Seller.objects.filter(cash_seller__active=True,
                                           cash_seller__vendor=request.user.vendor.vendor).distinct()

        seller = seller.filter(name__icontains=word)
    f = CashbekFilter(request.GET, queryset=cashbek)
    cashbek = f.qs
    seller_list = []
    incoming = 0
    expening = 0
    for sel in seller:
        incom = cashbek.filter(types=1, seller=sel).count()
        expen = cashbek.filter(types=2, seller=sel).count()
        incoming += cashbek.filter(types=1, seller=sel).aggregate(all_price=Sum(F("price")))["all_price"] if \
            cashbek.filter(types=1, seller=sel).aggregate(all_price=Sum(F("price")))["all_price"] else 0
        expening += cashbek.filter(types=2, seller=sel).aggregate(all_price=Sum(F("price")))["all_price"] if \
            cashbek.filter(types=2, seller=sel).aggregate(all_price=Sum(F("price")))["all_price"] else 0
        seller_list.append({"pk": sel.pk, "name": sel.name, "count": incom + expen, "incom": incom, "expen": expen})
    seller_list = sorted(seller_list, key=lambda x: x['count'], reverse=True)
    pagination = Paginator(seller_list, 10)
    page_number = request.GET.get('page')
    page_obj = pagination.get_page(page_number)
    context.update({"obj": page_obj, "filter": f, 'incoming': incoming, 'expening': expening})
    return render(request, "shop.html", context)


@user_passes_test(dashboard_access, login_url="signin")
def shop_detail(request, pk):
    context = {"shop": True}
    seller = Seller.objects.get(pk=pk)
    if request.user.is_manager():
        cashbek = Cashbek.objects.filter(active=True, seller=seller)
    else:
        cashbek = Cashbek.objects.filter(vendor=request.user.vendor.vendor, active=True, seller=seller)
    f = CashbekFilter(request.GET, queryset=cashbek)
    cashbek = f.qs.order_by("-created_at")
    if request.GET.get('start_date'):
        context.update({"start_date": request.GET.get('start_date')})
    if request.GET.get('end_date'):
        context.update({"end_date": request.GET.get('end_date')})
    if request.GET.get('q'):
        word = request.GET.get('q')
        if request.user.is_manager():
            cashbek = cashbek.filter(Q(product__model__icontains=word) | Q(product__imei1__icontains=word) | Q(
                promo__name__icontains=word) | Q(user_phone__icontains=word) | Q(user__last_name__icontains=word))
        else:
            cashbek = cashbek.filter(Q(product__model__icontains=word) | Q(product__imei1__icontains=word) | Q(
                promo__name__icontains=word))
    incom = cashbek.filter(types=1).aggregate(all_price=Sum(F("price")))["all_price"]
    excom = cashbek.filter(types=2).aggregate(all_price=Sum(F("price")))["all_price"]
    context.update({"filter": f, 'incom': incom, 'excom': excom, 'seller': seller})
    pagination = Paginator(cashbek, 10)
    page_number = request.GET.get('page')
    page_obj = pagination.get_page(page_number)
    context.update({"obj": page_obj, "filter": f})
    return render(request, "shop_detail.html", context)


@user_passes_test(dashboard_access, login_url="signin")
def seller_aggrement(request):
    context = {"seller_aggrement": True}
    if request.user.is_manager():
        cashbek = Cashbek.objects.filter(active=True)
        payment = PaymentForSeller.objects.all()
        total = cashbek.filter(types=2).aggregate(all_price=Sum(F("price")))["all_price"] if \
            cashbek.filter(types=2).aggregate(all_price=Sum(F("price")))["all_price"] else 0
        paid = payment.aggregate(all_price=Sum(F("amount")))["all_price"] if \
            payment.aggregate(all_price=Sum(F("amount")))["all_price"] else 0
        residual = total - paid
        seller = Seller.objects.all()
        if request.GET.get('q'):
            word = request.GET.get('q')
            seller = seller.filter(name__icontains=word)
        pagination = Paginator(seller, 10)
        page_number = request.GET.get('page')
        page_obj = pagination.get_page(page_number)
        context.update({'total': total, 'paid': paid, 'residual': residual, 'obj': page_obj})
        return render(request, "aggrement.html", context)
    else:
        return redirect("home")


@user_passes_test(dashboard_access, login_url="signin")
def seller_aggrement_detail(request, pk):
    context = {"seller_aggrement": True}
    if request.user.is_manager():
        if request.POST:
            seller = Seller.objects.get(pk=pk)
            PaymentForSeller.objects.create(seller=seller, amount=request.POST.get("amount"),
                                            descriptions=request.POST.get("comment"))
            return redirect("seller_paid_detail", seller.pk)
        seller = Seller.objects.get(pk=pk)
        payments = PaymentForSeller.objects.filter(seller=seller).order_by('-created_at')
        f = PaymentSeller(request.GET, queryset=payments)
        payments = f.qs.order_by("-id")
        if request.GET.get('start_date'):
            context.update({"start_date": request.GET.get('start_date')})
        if request.GET.get('end_date'):
            context.update({"end_date": request.GET.get('end_date')})
        pagination = Paginator(payments, 10)
        page_number = request.GET.get('page')
        page_obj = pagination.get_page(page_number)
        context.update({"obj": page_obj, 'seller': seller, 'filter': f})
        return render(request, "aggrement_detail.html", context)
    else:
        return redirect("home")


@user_passes_test(dashboard_access, login_url="signin")
def vendor_aggrement(request):
    context = {"vendor_aggrement": True}
    if request.user.is_manager():
        cashbek = Cashbek.objects.filter(active=True)
        payment = PaymentOfVendor.objects.all()
        total = cashbek.filter(types=1).aggregate(all_price=Sum(F("price")))["all_price"] if \
            cashbek.filter(types=1).aggregate(all_price=Sum(F("price")))["all_price"] else 0
        paid = payment.aggregate(all_price=Sum(F("amount")))["all_price"] if \
            payment.aggregate(all_price=Sum(F("amount")))["all_price"] else 0
        residual = total - paid
        vendor = Vendor.objects.all()
        if request.GET.get('q'):
            word = request.GET.get('q')
            vendor = vendor.filter(name__icontains=word)
        pagination = Paginator(vendor, 10)
        page_number = request.GET.get('page')
        page_obj = pagination.get_page(page_number)
        context.update({'total': total, 'paid': paid, 'residual': residual, 'obj': page_obj})
        return render(request, "aggrement.html", context)
    else:
        return redirect("home")


@user_passes_test(dashboard_access, login_url="signin")
def vendor_aggrement_detail(request, pk):
    context = {"vendor_aggrement": True}
    if request.user.is_manager():
        if request.POST:
            vendor = Vendor.objects.get(pk=pk)
            PaymentOfVendor.objects.create(vendor=vendor, amount=request.POST.get("amount"),
                                           descriptions=request.POST.get("comment"))
            return redirect("vendor_paid_detail", vendor.pk)
        vendor = Vendor.objects.get(pk=pk)
        payments = PaymentOfVendor.objects.filter(vendor=vendor).order_by('-created_at')
        f = PaymentVendor(request.GET, queryset=payments)
        payments = f.qs.order_by("-id")
        if request.GET.get('start_date'):
            context.update({"start_date": request.GET.get('start_date')})
        if request.GET.get('end_date'):
            context.update({"end_date": request.GET.get('end_date')})
        pagination = Paginator(payments, 10)
        page_number = request.GET.get('page')
        page_obj = pagination.get_page(page_number)
        context.update({"obj": page_obj, 'vendor': vendor, 'filter': f})
        return render(request, "aggrement_detail.html", context)
    else:
        return redirect("home")


@user_passes_test(dashboard_access, login_url="signin")
def vendor_detail(request):
    context = {"vendor_aggrement": True}
    if not request.user.is_manager():
        vendor = request.user.vendor.vendor
        payments = PaymentOfVendor.objects.filter(vendor=vendor).order_by('-created_at')
        f = PaymentVendor(request.GET, queryset=payments)
        payments = f.qs.order_by("-id")
        if request.GET.get('start_date'):
            context.update({"start_date": request.GET.get('start_date')})
        if request.GET.get('end_date'):
            context.update({"end_date": request.GET.get('end_date')})
        pagination = Paginator(payments, 10)
        page_number = request.GET.get('page')
        page_obj = pagination.get_page(page_number)
        context.update({"obj": page_obj, 'vendor': vendor, 'filter': f})
        return render(request, "aggrement_detail.html", context)
    else:
        return redirect("home")


@user_passes_test(dashboard_access, login_url="signin")
def export_cashbek(request):
    if request.user.is_manager():
        cashbek = Cashbek.objects.filter(active=True).order_by("-created_at")
        if request.POST.get('vendor_id'):
            cashbek = Cashbek.objects.filter(active=True, vendor=Vendor.objects.get(pk=request.POST.get('vendor_id'))).order_by("-created_at")
        file_location = BASE_DIR / 'file/cashbek_info.xlsx'
        file_send = BASE_DIR / f'file/cashbek.xlsx'
        wb = load_workbook(file_location)
        sheet = wb.active
        for i_row, cash in enumerate(cashbek, start=4):
            product = cash.product
            seller = cash.seller
            user = cash.user if cash.user else 0
            sheet.cell(row=i_row, column=1, value=get_tashkent_time(cash.created_at).strftime('%d-%m-%Y %H:%M'))
            sheet.cell(row=i_row, column=2, value=product.model)
            sheet.cell(row=i_row, column=3, value=product.imei1)
            sheet.cell(row=i_row, column=4, value=product.sku)
            sheet.cell(row=i_row, column=5, value=cash.promo.name if cash.promo else "")
            sheet.cell(row=i_row, column=6, value=cash.amount)
            sheet.cell(row=i_row, column=7, value=cash.get_types_display())
            sheet.cell(row=i_row, column=8, value=cash.vendor.name)
            sheet.cell(row=i_row, column=9, value=seller.name)
            sheet.cell(row=i_row, column=10, value=seller.get_region_display())
            sheet.cell(row=i_row, column=11, value=seller.get_district_display())
            sheet.cell(row=i_row, column=12, value=seller.seller.phone)
            if user:
                sheet.cell(row=i_row, column=13, value=user.first_name)
                sheet.cell(row=i_row, column=14, value=user.last_name)
                sheet.cell(row=i_row, column=15, value=user.passport_number)
                sheet.cell(row=i_row, column=16, value=user.pinfl)
                sheet.cell(row=i_row, column=17, value=user.citizenship)
                sheet.cell(row=i_row, column=18, value=user.birth_date.strftime('%d-%m-%Y'))
                sheet.cell(row=i_row, column=19, value=user.birth_place)
                sheet.cell(row=i_row, column=20, value=user.gender)
                sheet.cell(row=i_row, column=21, value=user.region)
                sheet.cell(row=i_row, column=22, value=user.district)
                sheet.cell(row=i_row, column=23, value=user.address)
                sheet.cell(row=i_row, column=24, value=cash.user_phone)
            else:
                sheet.cell(row=i_row, column=13, value="Удаленный аккаунт")
                sheet.cell(row=i_row, column=24, value=cash.user_phone)
        wb.save(file_send)
        with open(file_send, 'rb') as f:
            file_data = f.read()
        response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
        response[
            'Content-Disposition'] = f'attachment; filename="cashbek_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'
        os.remove(file_send)
        return response
    else:
        cashbek = Cashbek.objects.filter(active=True, vendor=request.user.vendor.vendor).order_by("-created_at")
        if request.POST.get('vendor_id'):
            cashbek = Cashbek.objects.filter(active=True, vendor=Vendor.objects.get(pk=request.POST.get('vendor_id')))
        file_location = BASE_DIR / 'file/cashbek_info1.xlsx'
        file_send = BASE_DIR / f'file/cashbek.xlsx'
        wb = load_workbook(file_location)
        sheet = wb.active
        for i_row, cash in enumerate(cashbek, start=4):
            product = cash.product
            seller = cash.seller
            sheet.cell(row=i_row, column=1, value=get_tashkent_time(cash.created_at).strftime('%d-%m-%Y %H:%M'))
            sheet.cell(row=i_row, column=2, value=product.model)
            sheet.cell(row=i_row, column=3, value=product.imei1)
            sheet.cell(row=i_row, column=4, value=product.sku)
            sheet.cell(row=i_row, column=5, value=cash.promo.name if cash.promo else "")
            sheet.cell(row=i_row, column=6, value=cash.amount)
            sheet.cell(row=i_row, column=7, value=cash.get_types_display())
            sheet.cell(row=i_row, column=8, value=cash.vendor.name)
            sheet.cell(row=i_row, column=9, value=seller.name)
            sheet.cell(row=i_row, column=10, value=seller.get_region_display())
            sheet.cell(row=i_row, column=11, value=seller.get_district_display())
            sheet.cell(row=i_row, column=12, value=seller.seller.phone)
        wb.save(file_send)
        with open(file_send, 'rb') as f:
            file_data = f.read()
        response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
        response[
            'Content-Disposition'] = f'attachment; filename="cashbek_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'
        os.remove(file_send)
        return response


@user_passes_test(dashboard_access, login_url="signin")
def status_cashbek(request):
    if request.POST.get("active"):
        Cashbek.objects.filter(pk=request.POST.get("cashbek_id")).update(active=False,
                                                                         description=request.POST.get("comment"))
        messages.success(request, "Кэшбэк успешно деактивирован")
    elif request.POST.get("inactive"):
        Cashbek.objects.filter(pk=request.POST.get("cashbek_id")).update(active=True)
        messages.success(request, "Кэшбэк успешно активирован")
    return redirect(request.META.get('HTTP_REFERER'))






