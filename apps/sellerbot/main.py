import datetime
import os

from django.db.models import Sum, F
from django.utils import timezone
from openpyxl.reader.excel import load_workbook
from telegram import Update, InputFile
from telegram.ext import CallbackContext
import qrcode
from io import BytesIO
from apps.main.models import Products, ACTIVE, PriceProduct, QR_code, INCOME, Token_confirm, EXPENSE, Cashbek, \
    PaymentForSeller
from apps.main.utility import get_tashkent_time
from apps.sellerbot.buttons import *
from apps.sellerbot.models import SellerTemp
from apps.users.models import Seller
from config.settings import BASE_DIR


def start_handler(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id
    try:
        seller = Seller.objects.get(telegram_id=user_id, seller__is_active=True)
        text = f"<strong>Продавец: </strong>{seller.name}\n<strong>ID: </strong><code>{seller.telegram_id}</code>"
        update.message.reply_html(text, reply_markup=main_buttons)
        try:
            obj = SellerTemp.objects.get(tg_id=user_id)
            obj.step = 0
            obj.save()
        except:
            SellerTemp.objects.create(tg_id=user_id, step=0)
    except Exception as e:
        print(e)


def core(update: Update, context: CallbackContext):
    user_id = str(update.message.from_user.id)
    try:
        seller = Seller.objects.get(telegram_id=user_id, seller__is_active=True)
        msg = update.message.text
        step = SellerTemp.objects.get(tg_id=user_id)
        if step.step == 0 and msg == 'Начисление Кэшбэка':
            SellerTemp.objects.filter(tg_id=user_id).update(step=1, qr_type=1)
            update.message.reply_html("Введите IMEI или серийный номер", reply_markup=home_button)
        elif step.step == 0 and msg == 'Cписание Кэшбэка':
            SellerTemp.objects.filter(tg_id=user_id).update(step=1, qr_type=2)
            update.message.reply_html("Введите IMEI или серийный номер", reply_markup=home_button)
        elif step.step == 1 and msg != '🏠Главный страница':
            SellerTemp.objects.filter(tg_id=user_id).update(step=2, imei=msg)
            if step.qr_type == 1:
                objs = Products.objects.filter(imei1=msg, is_active=True)
                if objs.exists():
                    today = timezone.now()
                    promo = objs.first().promo.filter(start__lte=today, end__gte=today, status=ACTIVE)
                    if promo.exists():
                        qr_obj = QR_code.objects.filter(product=objs.first(), is_used=False, expiry_date__gt=today)
                        if qr_obj.exists():
                            SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                            update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                                      f"</strong><code>{objs.first().imei1}</code>\n\n"
                                                      "<i>❌Есть активный QR-код</i>",
                                                      reply_markup=main_buttons)
                        else:
                            SellerTemp.objects.filter(tg_id=user_id).update(step=2, imei=msg)
                            update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n"
                                                      f"<strong>IMEI: </strong><code>{objs.first().imei1}</code>\n"
                                                      f"<strong>SKU: </strong>{objs.first().sku}\n"
                                                      "<strong>Сумма кэшбэка: </strong>" + "{:,}".format(
                                PriceProduct.objects.get(product=objs.first(),
                                                         promo=promo.first()).price) + " Cум\n"
                                                                                       f"<strong>Тип: </strong>Начисление Кэшбэка\n\n"
                                                                                       "Продукт и акция активны в данный момент. Нажмите на кнопку ниже, чтобы воспользоваться акцией👇",
                                                      reply_markup=qr_button)
                    else:
                        SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                        update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                                  f"</strong><code>{objs.first().imei1}</code>\n\n"
                                                  "<i>❌Этот товар недоступен ни в одной акции</i>",
                                                  reply_markup=main_buttons)
                else:
                    SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                    update.message.reply_html("❌Нет активного продукта с таким IMEI или серийным номером",
                                              reply_markup=main_buttons)

            elif step.qr_type == 2:
                objs = Products.objects.filter(imei1=msg, is_active=True)
                if objs.exists():
                    today = timezone.now()
                    promo = objs.first().promo.filter(start__lte=today, end__gte=today, status=ACTIVE)
                    qr_obj = QR_code.objects.filter(product=objs.first(), is_used=False, expiry_date__gt=today)
                    if qr_obj.exists():
                        SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                        update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                                  f"</strong><code>{objs.first().imei1}</code>\n\n"
                                                  "<i>❌Есть активный QR-код</i>",
                                                  reply_markup=main_buttons)
                    else:
                        SellerTemp.objects.filter(tg_id=user_id).update(step=2, imei=msg)
                        update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n"
                                                  f"<strong>IMEI: </strong><code>{objs.first().imei1}</code>\n"
                                                  f"<strong>SKU: </strong>{objs.first().sku}\n"
                                                  "<strong>Сумма кэшбэка: </strong>" + "{:,}".format(
                            PriceProduct.objects.get(product=objs.first(),
                                                     promo=promo.first()).price if promo.exists() else 0) + " Cум\n"
                                                                                                            f"<strong>Тип: </strong>Cписание Кэшбэка\n\n"
                                                                                                            "Продукт активны в данный момент. Нажмите кнопку ниже, чтобы получить кэшбэк👇",
                                                  reply_markup=qr_button)
                else:
                    SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                    update.message.reply_html("❌Нет активного продукта с таким IMEI или серийным номером",
                                              reply_markup=main_buttons)

        elif step.step == 2 and msg == 'Генерация QR-кода':
            if step.qr_type == 1:
                objs = Products.objects.filter(imei1=step.imei, is_active=True)
                if objs.exists():
                    today = timezone.now()
                    promo = objs.first().promo.filter(start__lte=today, end__gte=today, status=ACTIVE)
                    if promo.exists():
                        qr_obj = QR_code.objects.filter(product=objs.first(), is_used=False, expiry_date__gt=today)
                        if qr_obj.exists():
                            SellerTemp.objects.filter(tg_id=user_id).update(step=0, imei=msg)
                            update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                                      f"</strong><code>{objs.first().imei1}</code>\n\n"
                                                      "<i>❌Есть активный QR-код</i>",
                                                      reply_markup=main_buttons)
                        else:
                            if Token_confirm.objects.filter(product=objs.first(), expiry_date__gt=today).exists():
                                SellerTemp.objects.filter(tg_id=user_id).update(step=0, imei=msg)
                                update.message.reply_html(
                                    f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                    f"</strong><code>{objs.first().imei1}</code>\n\n"
                                    "<i>❌Товар находится в процессе покупки в другом сеансе</i>",
                                    reply_markup=main_buttons)
                            else:
                                qr_code = QR_code.objects.create(product=objs.first(), types=INCOME,
                                                                 seller=Seller.objects.get(telegram_id=user_id))
                                qr = qrcode.QRCode(
                                    version=1,
                                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                                    box_size=10,
                                    border=4,
                                )
                                qr.add_data(qr_code.qr_id)
                                qr.make(fit=True)
                                img_stream = BytesIO()
                                qr.make_image(fill_color="green", back_color="white").save(img_stream)
                                binary_data = img_stream.getvalue()
                                message = update.message.reply_photo(photo=InputFile(binary_data),
                                                                     caption="Тип <strong>Начисление Кэшбэка</strong>\n"
                                                                             "Срок действия <strong>2 минуты</strong>\n",
                                                                     parse_mode="HTML")
                                qr_code.message_id = message["message_id"]
                                qr_code.chat_id = user_id
                                qr_code.save_edit()
                                SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                                update.message.reply_html("Главный страница", reply_markup=main_buttons)

                    else:
                        SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                        update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                                  f"</strong><code>{objs.first().imei1}</code>\n\n"
                                                  "<i>❌Этот товар недоступен ни в одной акции</i>",
                                                  reply_markup=main_buttons)
                else:
                    SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                    update.message.reply_html("❌Нет активного продукта с таким IMEI или серийным номером",
                                              reply_markup=main_buttons)
            elif step.qr_type == 2:
                objs = Products.objects.filter(imei1=step.imei, is_active=True)
                if objs.exists():
                    today = timezone.now()
                    qr_obj = QR_code.objects.filter(product=objs.first(), is_used=False, expiry_date__gt=today)
                    if qr_obj.exists():
                        SellerTemp.objects.filter(tg_id=user_id).update(step=0, imei=msg)
                        update.message.reply_html(f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                                  f"</strong><code>{objs.first().imei1}</code>\n\n"
                                                  "<i>❌Есть активный QR-код</i>",
                                                  reply_markup=main_buttons)
                    else:
                        if Token_confirm.objects.filter(product=objs.first(), expiry_date__gt=today).exists():
                            SellerTemp.objects.filter(tg_id=user_id).update(step=0, imei=msg)
                            update.message.reply_html(
                                f"<strong>Модел: </strong>{objs.first().model}\n<strong>IMEI: "
                                f"</strong><code>{objs.first().imei1}</code>\n\n"
                                "<i>❌Товар находится в процессе покупки в другом сеансе</i>",
                                reply_markup=main_buttons)
                        else:
                            qr_code = QR_code.objects.create(product=objs.first(), types=EXPENSE,
                                                             seller=Seller.objects.get(telegram_id=user_id))
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_L,
                                box_size=10,
                                border=4,
                            )
                            qr.add_data(qr_code.qr_id)
                            qr.make(fit=True)
                            img_stream = BytesIO()
                            qr.make_image(fill_color="red", back_color="white").save(img_stream)
                            binary_data = img_stream.getvalue()
                            message = update.message.reply_photo(photo=InputFile(binary_data),
                                                                 caption="Тип <strong>Cписание Кэшбэка</strong>\n"
                                                                         "Срок действия <strong>2 минуты</strong>\n",
                                                                 parse_mode="HTML")
                            qr_code.message_id = message["message_id"]
                            qr_code.chat_id = user_id
                            qr_code.save_edit()
                            SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                            update.message.reply_html("Главный страница", reply_markup=main_buttons)


                else:
                    SellerTemp.objects.filter(tg_id=user_id).update(step=0)
                    update.message.reply_html("❌Нет активного продукта с таким IMEI или серийным номером",
                                              reply_markup=main_buttons)

        elif step.step == 0 and msg == 'Баланс':
            cashbek = Cashbek.objects.filter(seller=seller, active=True, types=2)
            payment = PaymentForSeller.objects.filter(seller=seller)
            total = cashbek.aggregate(all_price=Sum(F("amount")))["all_price"] if \
                cashbek.aggregate(all_price=Sum(F("amount")))["all_price"] else 0
            paid = payment.aggregate(all_price=Sum(F("amount")))["all_price"] if \
                payment.aggregate(all_price=Sum(F("amount")))["all_price"] else 0
            update.message.reply_html("<strong>⬆️Сумма Cписание Кэшбэка: </strong>{:,} Сум\n<strong>⬇️Оплаченный: </strong>{:,} Сум\n\n<strong>↔️Баланс: </strong>{:,} Сум".format(total, paid, total - paid))

        elif step.step == 0 and msg == 'Кэшбэки':
            SellerTemp.objects.filter(tg_id=user_id).update(step=3)
            update.message.reply_html("Выбирать👇", reply_markup=report_button)

        elif step.step == 3 and msg == 'Посмотреть последние кэшбэки':
            cashbek = Cashbek.objects.filter(seller=seller, active=True).order_by('-created_at')[:5]
            SellerTemp.objects.filter(tg_id=user_id).update(step=0)
            for i in cashbek:
                text = f"<strong>Модел: </strong>{i.product.model}\n"
                text += f"<strong>IMEI: </strong><code>{i.product.imei1}</code>\n"
                text += f"<strong>Дата: </strong>{get_tashkent_time(i.created_at).strftime('%d-%m-%Y %H:%M')}\n"
                text += "<strong>Сумма кэшбэка: </strong>{:,} Сум\n".format(i.amount)
                text += f"<strong>Клиент: </strong>{i.user.first_name[0]}.{i.user.last_name}({i.user.simple_user.phone})"
                update.message.reply_html(text)
            update.message.reply_html("Выбирать👇", reply_markup=main_buttons)

        elif step.step == 3 and msg == 'Полный список кэшбэков и выплат':
            cashbek = Cashbek.objects.filter(seller=seller, active=True).order_by("-created_at")
            payment = PaymentForSeller.objects.filter(seller=seller).order_by("-created_at")
            total = cashbek.aggregate(all_price=Sum(F("amount")))["all_price"] if \
                cashbek.aggregate(all_price=Sum(F("amount")))["all_price"] else 0
            paid = payment.aggregate(all_price=Sum(F("amount")))["all_price"] if \
                payment.aggregate(all_price=Sum(F("amount")))["all_price"] else 0

            SellerTemp.objects.filter(tg_id=user_id).update(step=0)
            file_location = BASE_DIR / 'file/report_seller.xlsx'
            file_send = BASE_DIR / f'file/{seller.name}_report.xlsx'
            wb = load_workbook(file_location)
            sheet = wb['cashbek']
            for i_row, cash in enumerate(cashbek, start=2):
                sheet.cell(row=i_row, column=1, value=get_tashkent_time(cash.created_at).strftime('%d-%m-%Y %H:%M'))
                sheet.cell(row=i_row, column=2, value=cash.product.model)
                sheet.cell(row=i_row, column=3, value=cash.product.imei1)
                sheet.cell(row=i_row, column=4, value=cash.amount)
                sheet.cell(row=i_row, column=5, value=f"{cash.user.first_name[0]}.{cash.user.last_name}({cash.user.simple_user.phone})")
                sheet.cell(row=i_row, column=6, value=cash.get_types_display())
            payment_sheet = wb["payment"]
            payment_sheet["C2"].value = total
            payment_sheet["C3"].value = paid
            payment_sheet["C4"].value = total - paid
            for i_row, cash in enumerate(payment, start=7):
                payment_sheet.cell(row=i_row, column=1, value=get_tashkent_time(cash.created_at).strftime('%d-%m-%Y %H:%M'))
                payment_sheet.cell(row=i_row, column=2, value=cash.amount)
                payment_sheet.cell(row=i_row, column=3, value=cash.descriptions)
            wb.save(file_send)
            update.message.reply_document(open(file_send, 'rb'), reply_markup=main_buttons)
            os.remove(file_send)
        elif msg == '🏠Главный страница':
            SellerTemp.objects.filter(tg_id=user_id).update(step=0)
            update.message.reply_html("Главный страница", reply_markup=main_buttons)
    except Exception as e:
        print(e)
